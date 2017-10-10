
import os, sys
import openpyxl
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

class OpenPyExcel(object):
    """
    Read/Write Excel
    The API just can read/write xlsx file. 
    """
    def __init__(self, filename, **args):
        super(OpenPyExcel, self).__init__()
        self.filename = filename
        self._workbook = None
        self._active_sheet = None
        
        
    def load_workbook(self, readonly=False):
        if not os.path.isfile(self.filename):
            if readonly is True:
                sys.stderr.write( "OpenPyExcel Error: \"%s\" No such file.\n"%(self.filename) )
                self._workbook = None
                return False
            else:
                try:
                    self._workbook = openpyxl.Workbook()
                    return True
                except:
                    return False
        try:
            self._workbook = openpyxl.load_workbook(self.filename, readonly)
            return True
        except:
            sys.stderr.write( "OpenPyExcel Error: Load \"%s\" failure.\n"%(self.filename) )
            self._workbook = None
            return False
        
        
    @property
    def is_opened(self):
        if self._workbook is not None:
            return True
        else:
            return False
        
   
    #@property     
    #def active_sheet(self):
    #    return self._active_sheet
    
    
    #@active_sheet.setter
    #def active_sheet(self, sheet_name):
    #    self.open_sheet(sheet_name)

    
    @property
    def rows(self):
        return self._active_sheet.max_row


    @property
    def columns(self):
        return self._active_sheet.max_column
    
    
    def open_sheet(self, sheet_name):
        if self._workbook is None:
            self._active_sheet = None
            return False
        try:
            ws = self._workbook.get_sheet_by_name(sheet_name)
        except:
            self._active_sheet = None
            return False
        self._workbook.active = self._workbook.get_index(ws)
        self._active_sheet = self._workbook.active
        return True
    
    
    def add_sheet(self, sheet_name, option='w'):
        """
        option: r, w
        r - read only
        w - delete the original data & sheet, then add a blank sheet
        """
        if option == 'w':
            if not self.del_sheet(sheet_name):
                return False
        if self._is_sheet_in_workbook(sheet_name) is False:
            new_sheet = self._workbook.create_sheet(title = sheet_name)
            if new_sheet is not None:
                return True
        return False

                
    def del_sheet(self, sheet_name):
        if self._is_sheet_in_workbook(sheet_name) is True:
            ws = self._workbook.get_sheet_by_name(sheet_name)
            #self._workbook.remove_sheet(ws)
            self._workbook.remove(ws)
            if self._is_sheet_in_workbook(sheet_name):
                return False
        return True
    
    
    def rename_sheet(self, old_sheet_name, new_sheet_name):
        if self._is_sheet_in_workbook(old_sheet_name) is True:
            ws = self._workbook.get_sheet_by_name(old_sheet_name)
            ws.title = new_sheet_name
            if self._is_sheet_in_workbook(new_sheet_name) is True: 
                return True
        return False
    

    def get_rows(self):
        return self._active_sheet.max_row
    
    
    def get_columns(self):
        return self._active_sheet.max_column
        
        
    def get_sheet_names(self):
        """
        return list of the sheet names
        """
        return self._workbook.get_sheet_names()
    
    
    def get_cell_value(self, row, col):
        return self._active_sheet.cell(row=row, column=col).value


    def get_cell_value_by_coordinate(self, coordinate):
        """
        :param coordinate: coordinates of the cell (e.g. 'B12')
        :type coordinate: string
        """
        return self._active_sheet.cell(coordinate=coordinate).value
    
    
    def write_cell(self, row, col, data):
        return self._active_sheet.cell(row=row, column=col, value=data)
      
        
    def write_cell_by_coordinate(self, coordinate, data):
        """
        :param coordinate: coordinates of the cell (e.g. 'B12')
        :type coordinate: string
        """
        #return self._active_sheet.cell(coordinate=coordinate, value=data)
        self._active_sheet[coordinate].value = data 
        
        
    def write_cell_string(self, row, col, str_data=""):
        return self.write_cell(row, col, str_data)
        
        
    def write_cell_number(self, row, col, num=0):
        cell = self.write_cell(row, col, num)
        cell.number_format
    
    
    def close(self, save=False):
        if save:
            return self._save()
        else:
            return True
            
    def close_save(self):
        return self._save()
    
    def _save(self):
        try:
            self._workbook.save(filename = self.filename)
            return True
        except:
            return False
    
 
    def _is_sheet_in_workbook(self, ws_name):
        if ws_name in self.get_sheet_names():
            return True
        else:
            return False
        
    
    
    
    
        
